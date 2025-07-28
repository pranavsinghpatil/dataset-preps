import openpyxl
import os

# Construct the absolute paths
port_dir = 'D:\\repofixs\\port'
source_file = os.path.join(port_dir, 'finedata.xlsx')
output_file = os.path.join(port_dir, 'kfine.xlsx')

# Load the source workbook
source_workbook = openpyxl.load_workbook(source_file)
source_sheet = source_workbook.active

# Create a new workbook for the output
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# Find the index of the "Audio" column
header = [cell.value for cell in source_sheet[1]]
audio_column_index = -1
for i, h in enumerate(header):
    if h and 'audio' in h.lower():
        audio_column_index = i + 1
        break


if audio_column_index != -1:
    # Copy headers, skipping the "Audio" column
    new_header = [h for i, h in enumerate(header) if i + 1 != audio_column_index]
    output_sheet.append(new_header)

    # Copy data rows, skipping the "Audio" column
    for row in source_sheet.iter_rows(min_row=2):
        new_row = []
        for i, cell in enumerate(row):
            if i + 1 != audio_column_index:
                # Create a new cell in the output sheet and copy the value
                # and other properties like font, border, etc. if needed
                new_cell = output_sheet.cell(row=output_sheet.max_row + 1, column=len(new_row) + 1)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
        #This is a bit of a hack, but since we are creating cells one by one, we need to append an empty list to create a new row
        output_sheet.append([])


    # copy images
    for image in source_sheet._images:
        output_sheet.add_image(image)

else:
    print("Audio column not found!")


# Save the new workbook
output_workbook.save(output_file)
