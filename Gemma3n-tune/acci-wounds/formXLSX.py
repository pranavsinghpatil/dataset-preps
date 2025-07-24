import csv
import json
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

def parse_jsonl(jsonl_path):
    rows = []
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            obj = json.loads(line)
            filename = obj.get('filename', '')
            messages = obj.get('messages', [])
            user_msg = ''
            ai_msg = ''
            for m in messages:
                if m.get('role') == 'user':
                    user_msg = m.get('content', '')
                    if isinstance(user_msg, dict):
                        user_msg = str(user_msg)
                    user_msg = user_msg.replace('\n', ' ').replace('\r', ' ')
                elif m.get('role') == 'assistant':
                    ai_msg = m.get('content', '')
                    if isinstance(ai_msg, dict):
                        ai_msg = str(ai_msg)
                    ai_msg = ai_msg.replace('\n', ' ').replace('\r', ' ')
            text = f"[User: {user_msg} , AI: {ai_msg}]"
            rows.append((filename, '', text))
    return rows

def parse_csv(csv_path):
    rows = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader, None)  # Skip header if present
        for row in reader:
            if len(row) == 3:
                row[2] = row[2].replace('\n', ' ').replace('\r', ' ')
                rows.append((row[0], row[1], row[2]))
            elif len(row) == 2:
                row[1] = row[1].replace('\n', ' ').replace('\r', ' ')
                rows.append((row[0], '', row[1]))
            elif len(row) == 1:
                rows.append((row[0], '', ''))
    return rows

def create_excel(data, image_dir, out_path):
    wb = Workbook()
    ws = wb.active
    ws.append(['Image', 'Audio', 'Text'])
    temp_images = []
    for idx, (filename, audio, text) in enumerate(data, start=2):
        img_path = os.path.join(image_dir, filename)
        ws.cell(row=idx, column=2, value=audio)
        ws.cell(row=idx, column=3, value=text)
        if os.path.exists(img_path) and filename.lower().endswith(('.jpg', '.png')):
            pil_img = PILImage.open(img_path)
            pil_img.thumbnail((128, 128))
            temp_img_path = f'temp_{idx}.png'
            pil_img.save(temp_img_path)
            xl_img = XLImage(temp_img_path)
            ws.add_image(xl_img, f'A{idx}')
            temp_images.append(temp_img_path)
        else:
            ws.cell(row=idx, column=1, value=filename)
    wb.save(out_path)
    for temp_img_path in temp_images:
        if os.path.exists(temp_img_path):
            os.remove(temp_img_path)

if __name__ == '__main__':
    image_dir = 'd:\\repofixs\\port\\images'
    csv_data = parse_csv('d:\\repofixs\\port\\datatillnow.csv')
    jsonl_data = parse_jsonl('d:\\repofixs\\port\\dataset.jsonl')
    all_data = csv_data + jsonl_data
    create_excel(all_data, image_dir, 'd:\\repofixs\\port\\finedata.xlsx')
    print('Excel file with embedded images written to finedata.xlsx')